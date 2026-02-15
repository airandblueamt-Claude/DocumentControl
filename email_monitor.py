
#!/usr/bin/env python3
"""Document Control Email Monitor - Main entry point.

Connects to Gmail via IMAP, fetches forwarded emails,
classifies them, saves attachments, and logs to Excel.
"""

import logging
import os
import sys
from datetime import datetime
from logging.handlers import RotatingFileHandler

from openpyxl import Workbook, load_workbook

from email_interface.auth import create_auth_handler
from email_interface.attachment_handler import AttachmentHandler
from email_interface.config import load_config, resolve_path
from email_interface.imap_client import ImapEmailClient
from email_interface.message_processor import MessageProcessor
from email_interface.persistence import ProcessingTracker

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_HEADERS = [
    'Transmittal No', 'Date', 'From', 'To', 'CC', 'Subject',
    'Reference', 'Type', 'Discipline', 'Response Required',
    'Attachment Count', 'Attachment Folder', 'Message ID',
]


def setup_logging(log_config):
    """Configure rotating file logger + console output."""
    log_file = resolve_path(log_config.get('log_file', 'logs/email_interface.log'), BASE_DIR)
    log_level = getattr(logging, log_config.get('log_level', 'INFO').upper(), logging.INFO)

    os.makedirs(os.path.dirname(log_file), exist_ok=True)

    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s')

    file_handler = RotatingFileHandler(log_file, maxBytes=10 * 1024 * 1024, backupCount=5)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)

    root_logger = logging.getLogger()
    root_logger.setLevel(log_level)
    root_logger.addHandler(file_handler)
    root_logger.addHandler(console_handler)


def update_excel_log(processed_emails, excel_path):
    """Append processed email rows to the Excel correspondence log."""
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Correspondence Log'
        ws.append(EXCEL_HEADERS)

    for entry in processed_emails:
        to_str = '; '.join(
            r.get('email', '') if isinstance(r, dict) else str(r)
            for r in entry.get('to', [])
        )
        cc_str = '; '.join(
            r.get('email', '') if isinstance(r, dict) else str(r)
            for r in entry.get('cc', [])
        )
        date_str = ''
        if entry.get('date'):
            date_str = entry['date'].strftime('%Y-%m-%d %H:%M') if hasattr(entry['date'], 'strftime') else str(entry['date'])

        att_folder = ''
        saved = entry.get('attachments_saved', [])
        if saved:
            att_folder = os.path.dirname(saved[0]['path'])

        row = [
            entry.get('transmittal_no', ''),
            date_str,
            entry.get('sender', ''),
            to_str,
            cc_str,
            entry.get('subject', ''),
            '; '.join(entry.get('references', [])),
            entry.get('doc_type', ''),
            entry.get('discipline', ''),
            'Y' if entry.get('response_required') else 'N',
            len(entry.get('attachments', [])),
            att_folder,
            entry.get('message_id', ''),
        ]
        ws.append(row)

    wb.save(excel_path)
    logging.getLogger(__name__).info("Excel log updated: %d rows added to %s", len(processed_emails), excel_path)


def process_mailbox():
    """Main workflow: fetch via IMAP, classify, save, log."""
    logger = logging.getLogger(__name__)

    # 1. Load configurations
    email_cfg = load_config(resolve_path('config/email_config.yaml', BASE_DIR))
    class_cfg = load_config(resolve_path('config/classification_config.yaml', BASE_DIR))

    setup_logging(email_cfg.get('logging', {}))
    logger.info("=" * 60)
    logger.info("Starting email processing run at %s", datetime.now().isoformat())

    # 2. Initialize components
    auth = create_auth_handler(email_cfg['auth'], email_cfg['mailbox']['email_address'])

    imap_client = ImapEmailClient(
        auth_handler=auth,
        host=email_cfg['server']['imap_host'],
        port=email_cfg['server'].get('imap_port', 993),
    )

    processor = MessageProcessor(class_cfg)

    att_cfg = email_cfg.get('attachments', {})
    attachment_handler = AttachmentHandler(
        base_path=resolve_path(att_cfg.get('base_path', 'Attachments'), BASE_DIR),
        max_size_mb=att_cfg.get('max_size_mb', 25),
        allowed_extensions=att_cfg.get('allowed_extensions'),
    )

    db_path = resolve_path('data/tracking.db', BASE_DIR)
    prefix = class_cfg.get('transmittal', {}).get('prefix', 'TRN')
    tracker = ProcessingTracker(db_path=db_path, prefix=prefix)

    excel_path = resolve_path(email_cfg.get('excel', {}).get('output_path', 'logs/correspondence_log.xlsx'), BASE_DIR)
    max_messages = email_cfg.get('polling', {}).get('max_messages_per_run', 50)

    try:
        # 3. Connect to mailbox
        imap_client.connect()
        folder = email_cfg.get('folders', {}).get('inbox', 'INBOX')
        imap_client.select_folder(folder)

        # 4. Search unread messages
        msg_ids = imap_client.search_unread()
        if not msg_ids:
            logger.info("No unread messages found")
            return

        msg_ids = msg_ids[:max_messages]
        logger.info("Processing up to %d messages", len(msg_ids))

        # 5. Process each message
        processed_emails = []
        for msg_id in msg_ids:
            try:
                raw_msg = imap_client.fetch_message(msg_id)
                msg_data = processor.parse_message(raw_msg)
                message_id = msg_data.get('message_id', '')

                # Skip if already processed
                if tracker.is_processed(message_id):
                    logger.info("Skipping already-processed: %s", message_id)
                    imap_client.mark_as_read(msg_id)
                    continue

                # Check skip filters
                should_process, reason = processor.should_process(msg_data)
                if not should_process:
                    logger.warning("Skipping: %s - %s", msg_data.get('subject', ''), reason)
                    imap_client.mark_as_read(msg_id)
                    continue

                # Classify
                msg_data['references'] = processor.extract_reference(msg_data)
                msg_data['doc_type'] = processor.classify_type(msg_data)
                msg_data['discipline'] = processor.classify_discipline(msg_data)
                msg_data['response_required'] = processor.detect_response_required(msg_data)

                # Generate transmittal number
                transmittal_no = tracker.get_next_transmittal_number(msg_data.get('date'))
                msg_data['transmittal_no'] = transmittal_no

                # Save attachments
                msg_data['attachments_saved'] = attachment_handler.save_attachments(msg_data, transmittal_no)

                processed_emails.append(msg_data)

                # Track as processed
                tracker.mark_processed(message_id, transmittal_no, msg_data)

                # Mark as read and move
                imap_client.mark_as_read(msg_id)
                dest_folder = email_cfg.get('folders', {}).get('processed', 'Processed')
                imap_client.move_to_folder(msg_id, dest_folder)

                logger.info("Processed: %s - %s", transmittal_no, msg_data.get('subject', ''))

            except Exception as e:
                logger.error("Failed to process message %s: %s", msg_id, e, exc_info=True)
                continue

        # 6. Update Excel log
        if processed_emails:
            update_excel_log(processed_emails, excel_path)

        logger.info("Run complete: %d emails processed", len(processed_emails))

    finally:
        tracker.close()
        imap_client.disconnect()


if __name__ == '__main__':
    process_mailbox()
